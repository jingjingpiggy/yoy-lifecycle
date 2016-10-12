from excel_general import *
import re
from openpyxl import Workbook, load_workbook, cell, utils

class excel_writer_pyxl(excel_writer):
    def open_file(self):
        excel_writer.open_file(self)

        print('Open excel file:', self.fileName)
        if self.newFile:
            self.wb = Workbook(guess_types=True)
        else:
            self.wb = load_workbook(self.fileName, guess_types=True)

        print('\tOpen excel file Done')
        return 0

    def write_excel(self, header, value_d, sheetname=''):
        print('Start write excel:', self.fileName)
        newSheet = True
        if self.newFile:
            ws = self.wb.create_sheet()
            if sheetname:
                ws.title = sheetname
        elif sheetname:
            if sheetname in self.wb.get_sheet_names():
                ws = self.wb[sheetname]
                newSheet = False
                print('\tWarning: openpyxl cannot fill existed sheet.Target sheet existed in file, will cover it!')
            else:
                ws = self.wb.create_sheet()
                ws.title = sheetname
        print('\tsheet:', ws.title)

        #delete old lines
        if not newSheet:
            for i in range(2, ws.max_row+1):
                for j in range(1, ws.max_column+1):
                    ws.cell(column=j, row=i).value = None

        #write first line
        if not newSheet:
            diff = False
            orig_header = [c.value for c in ws.rows[0] if c.value]
            if len(orig_header) != len(header):
                diff = True
            else:
                for j in range(1,len(header)+1):
                    if ws.cell(column=j, row=1).value != header[j-1]:
                        diff = True
            if diff:
                print('Warning: new headers diff from current! Will cover old with new.')
                print('\t original:', orig_header)
                print('\t new:', str(header))
        for j in range(1,len(header)+1):
            ws.cell(column=j, row=1).value = header[j-1]

        #write data
        for i in range(2,len(value_d)+2):
            line = value_d[i-2]
            for j in range(1,len(header)+1):
                v_write = line.get(header[j-1], '')
                try:
                    ws.cell(column=j, row=i).value = v_write
                except utils.exceptions.IllegalCharacterError:
                    #copied from openpyxl cell.py, which used to check IllegalCharacterError
                    ILLEGAL_CHARACTERS_p = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
                    v_write = ILLEGAL_CHARACTERS_p.sub(' ', v_write)
                    try:
                        ws.cell(column=j, row=i).value = v_write
                    except:
                        print('Error: exception in write excel! cell:', str(i)+'x'+str(j))
                        print('value:', v_write)
                        print('patch id:', value_d.get('number', 'fail to get num'))
        return 0

    def save_file(self):
        self.wb.save(filename = self.abs_fileName)
        print('\tsaved.')
        return 0

    def close_file(self):
        print('\tDone. Excel closed.(do nothing as using openpyxl)')
        return 0

    def quit(self):
        print('Quit excel for', self.fileName, '(do nothing as using openpyxl)')
        return 0
